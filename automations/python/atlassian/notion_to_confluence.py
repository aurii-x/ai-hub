#!/usr/bin/env python3
"""
Notion → Confluence Cloud Migration Script
Version: 1.0
Target DB: 06. Work (Notion) → WORK space (Confluence Cloud)

Architecture:
  Phase 0 — Build Sheet       : Query Notion, produce dry-run XLSX manifest
  Phase 1 — Create Pages      : Create Confluence pages with full property set
  Phase 2 — Copy Content      : Transfer blocks (text, rich text, images,
                                 attachments, embeds, PDFs, PPTX)
  Phase 3 — Lock Date Field   : Apply Page Properties macro + edit restriction
  Phase 4 — Mark Migrated     : Update build sheet status + write Jira backlink

All phases are idempotent. Re-running skips already-migrated pages.
Dry-run mode (DRY_RUN=True) never writes to Confluence or Jira.

Credentials — set as environment variables, never hardcode:
  NOTION_TOKEN          : Notion integration secret (Internal Integration Token)
  CONFLUENCE_BASE_URL   : e.g. https://yoursite.atlassian.net
  CONFLUENCE_EMAIL      : Atlassian account email
  CONFLUENCE_API_TOKEN  : Atlassian API token (from id.atlassian.com/manage-profile/security)
  JIRA_BASE_URL         : e.g. https://yoursite.atlassian.net  (usually same as Confluence)
  JIRA_EMAIL            : Jira account email (usually same as Confluence)
  JIRA_API_TOKEN        : Jira API token (usually same as Confluence)
  CONFLUENCE_SPACE_KEY  : Target Confluence space key, e.g. WORK
  CONFLUENCE_PARENT_ID  : (optional) Parent page ID in Confluence space root
"""

import os
import sys
import json
import base64
import logging
import mimetypes
import traceback
from datetime import datetime, timezone
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time

# Load credentials from .env file in same folder as this script
load_dotenv()

# ─────────────────────────────────────────────
# CONFIGURATION — only section you need to edit
# ─────────────────────────────────────────────

# === Source database ID (06. Work) ===
NOTION_DB_ID = "37761870-08a2-8199-941a-c6e481ccf03e"

# ── Mode ─────────────────────────────────────────────────────────────────────
# DRY_RUN = True  → builds XLSX manifest only, never writes to Confluence/Jira
# DRY_RUN = False → live migration runs in configurable batches
DRY_RUN = False

# ── Batch control ─────────────────────────────────────────────────────────────
# How many pages to migrate per run when DRY_RUN = False.
# Script stops after BATCH_SIZE pages so you can inspect results in Confluence
# before continuing. Set BATCH_SIZE = 0 to migrate ALL pending pages at once.
#
# Workflow:
#   1. Set DRY_RUN=False, BATCH_SIZE=5, BATCH_NUMBER=0 → run → check 5 pages
#   2. Happy with results → set BATCH_NUMBER=1 → run → check next 5
#   3. Keep incrementing BATCH_NUMBER until all rows show DONE in build sheet
#   4. Or set BATCH_SIZE=0 to run everything remaining in one shot
BATCH_SIZE   = 3   # pages per run  (0 = unlimited)
BATCH_NUMBER = 0   # which batch to run (0-indexed, auto-advances via build sheet)

# ── Build sheet path ──────────────────────────────────────────────────────────
BUILD_SHEET_PATH = "notion_confluence_build_sheet.xlsx"

# ── Jira backlink field ───────────────────────────────────────────────────────
# customfield_10085 = "Confluence Page(s)" confirmed in your Jira schema CSV
JIRA_CONFLUENCE_FIELD = "customfield_10085"

# ── Confluence target ─────────────────────────────────────────────────────────
CONFLUENCE_SPACE_KEY = os.getenv("CONFLUENCE_SPACE_KEY", "WORK")
CONFLUENCE_PARENT_ID = os.getenv("CONFLUENCE_PARENT_ID", "")  # blank = space root

# ── Rate limiting ─────────────────────────────────────────────────────────────
# Confluence Cloud sustained limit ~10 req/s. 0.15 s = ~6 req/s (safe)
REQUEST_DELAY_S = 0.15

# ─────────────────────────────────────────────
# PROPERTY MAPPING TABLE
# Notion property name → Confluence Page Properties macro key
# Derived from analysis of 06. Work schema + Jira schema
# Fields dropped: page_type, DB Name, formulas, rollups, relations,
#                 Notion Entity Type, ClickUp Task, Like, Rize*
# ─────────────────────────────────────────────
PROPERTY_MAP = {
    # Notion field name          : Confluence macro key
    "Name"                       : "title",           # handled as page title, not a macro key
    "Created"                    : "notion_created_date",  # write-once locked
    "Status"                     : "status",
    "Priority"                   : "priority",
    "Category"                   : "category",
    "Artifact Type"              : "artifact_type",
    "Tags"                       : "tags",
    "URL"                        : "source_url",
    "Company"                    : "company",
    "Project"                    : "project",
    "Date"                       : "entry_date",
    "Author"                     : "author",
    "AI summary"                 : "ai_summary",
    # Provenance (new — not in Notion schema, added for migration traceability)
    "_notion_page_id"            : "notion_page_id",
    "_jira_issue_link"           : "jira_issue_link",
    # Attachment fields handled separately (binary transfer, not macro key-value)
    # Dropped: Page Type, DB Name, formulas, rollups, all relations,
    #          Notion Entity Type, ClickUp Task, Like, Rize*, Since last change,
    #          Latest Child Created, Days Since Last Child Created, Children,
    #          Parent item, Sub-item, Updated (last_edited_time → not migrated)
}

# Notion status values → normalized label
STATUS_MAP = {
    "Not started" : "Not Started",
    "In progress" : "In Progress",
    "Done"        : "Done",
    "Archived"    : "Archived",
    "Backlog"     : "Backlog",
    "Deferred"    : "Deferred",
}

# ─────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("migration.log"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# API CLIENTS
# ─────────────────────────────────────────────

class NotionClient:
    """
    Wraps Notion REST API v1.
    Ref: https://developers.notion.com/reference
    All endpoints return parsed JSON dicts.
    """
    BASE = "https://api.notion.com/v1"
    VERSION = "2022-06-28"

    def __init__(self, token: str):
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization": f"Bearer {token}",
            "Notion-Version": self.VERSION,
            "Content-Type": "application/json",
        })

    def _get(self, path: str, params: dict = None) -> dict:
        time.sleep(REQUEST_DELAY_S)
        r = self.session.get(f"{self.BASE}{path}", params=params)
        r.raise_for_status()
        return r.json()

    def _post(self, path: str, body: dict) -> dict:
        time.sleep(REQUEST_DELAY_S)
        r = self.session.post(f"{self.BASE}{path}", json=body)
        r.raise_for_status()
        return r.json()

    def query_database(self, db_id: str) -> list:
        """Returns all pages in a database, handling pagination."""
        results, cursor = [], None
        while True:
            body = {"page_size": 100}
            if cursor:
                body["start_cursor"] = cursor
            data = self._post(f"/databases/{db_id}/query", body)
            results.extend(data.get("results", []))
            if not data.get("has_more"):
                break
            cursor = data.get("next_cursor")
        log.info(f"Notion: retrieved {len(results)} pages from DB {db_id}")
        return results

    def get_page_blocks(self, page_id: str) -> list:
        """
        Returns all block children of a page recursively.
        Handles pagination per block.
        Ref: https://developers.notion.com/reference/get-block-children
        """
        return self._get_children(page_id)

    def _get_children(self, block_id: str) -> list:
        results, cursor = [], None
        while True:
            params = {"page_size": 100}
            if cursor:
                params["start_cursor"] = cursor
            data = self._get(f"/blocks/{block_id}/children", params=params)
            blocks = data.get("results", [])
            for block in blocks:
                if block.get("has_children"):
                    block["_children"] = self._get_children(block["id"])
                else:
                    block["_children"] = []
            results.extend(blocks)
            if not data.get("has_more"):
                break
            cursor = data.get("next_cursor")
        return results

    def download_file(self, url: str) -> bytes:
        """Downloads a Notion file URL (signed S3 URL). Returns raw bytes."""
        time.sleep(REQUEST_DELAY_S)
        r = requests.get(url, timeout=60)
        r.raise_for_status()
        return r.content

    def extract_properties(self, page: dict) -> dict:
        """
        Extracts all properties from a Notion page into a flat dict.
        Handles: title, rich_text, select, multi_select, status, date,
                 created_time, last_edited_time, url, checkbox, files, people.
        Drops: relation, rollup, formula (Notion-internal, not portable).
        """
        props = {}
        raw = page.get("properties", {})

        def rich_text_to_str(rt_list: list) -> str:
            return "".join(rt.get("plain_text", "") for rt in rt_list)

        for name, prop in raw.items():
            t = prop.get("type")
            try:
                if t == "title":
                    props[name] = rich_text_to_str(prop.get("title", []))
                elif t == "rich_text":
                    props[name] = rich_text_to_str(prop.get("rich_text", []))
                elif t == "select":
                    s = prop.get("select")
                    props[name] = s.get("name", "") if s else ""
                elif t == "multi_select":
                    props[name] = ", ".join(
                        o.get("name", "") for o in prop.get("multi_select", [])
                    )
                elif t == "status":
                    s = prop.get("status")
                    raw_status = s.get("name", "") if s else ""
                    props[name] = STATUS_MAP.get(raw_status, raw_status)
                elif t == "date":
                    d = prop.get("date")
                    props[name] = d.get("start", "") if d else ""
                elif t == "created_time":
                    props[name] = prop.get("created_time", "")
                elif t == "last_edited_time":
                    props[name] = prop.get("last_edited_time", "")
                elif t == "url":
                    props[name] = prop.get("url") or ""
                elif t == "checkbox":
                    props[name] = str(prop.get("checkbox", False))
                elif t == "files":
                    files = prop.get("files", [])
                    urls = []
                    for f in files:
                        if f.get("type") == "external":
                            urls.append(f["external"]["url"])
                        elif f.get("type") == "file":
                            urls.append(f["file"]["url"])
                    props[name] = " | ".join(urls)
                elif t == "people":
                    people = prop.get("people", [])
                    props[name] = ", ".join(
                        p.get("name", p.get("id", "")) for p in people
                    )
                elif t in ("relation", "rollup", "formula"):
                    pass  # intentionally dropped
                else:
                    props[name] = ""
            except Exception as e:
                log.warning(f"Property extract error [{name}]: {e}")
                props[name] = ""

        # Inject provenance fields
        props["_notion_page_id"] = page.get("id", "")
        props["_notion_url"] = page.get("url", "")
        return props


class ConfluenceClient:
    """
    Wraps Confluence Cloud REST API v2 (content) and v1 (restrictions).
    Ref: https://developer.atlassian.com/cloud/confluence/rest/v2/intro/
         https://developer.atlassian.com/cloud/confluence/rest/v1/api-group-content-restrictions/
    """
    def __init__(self, base_url: str, email: str, token: str):
        self.base = base_url.rstrip("/")
        creds = base64.b64encode(f"{email}:{token}".encode()).decode()
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization": f"Basic {creds}",
            "Content-Type": "application/json",
            "Accept": "application/json",
        })

    def _v2(self, method: str, path: str, **kwargs) -> requests.Response:
        time.sleep(REQUEST_DELAY_S)
        url = f"{self.base}/wiki/api/v2{path}"
        r = self.session.request(method, url, **kwargs)
        if not r.ok:
            log.error(f"Confluence API {method} {url} → {r.status_code}: {r.text[:400]}")
        r.raise_for_status()
        return r

    def _v1(self, method: str, path: str, **kwargs) -> requests.Response:
        time.sleep(REQUEST_DELAY_S)
        url = f"{self.base}/wiki/rest/api{path}"
        r = self.session.request(method, url, **kwargs)
        if not r.ok:
            log.error(f"Confluence v1 API {method} {url} → {r.status_code}: {r.text[:400]}")
        r.raise_for_status()
        return r

    def find_page_by_title(self, space_key: str, title: str) -> dict | None:
        """
        Search for an existing page by exact title in a space.
        Ref: GET /wiki/api/v2/pages
        """
        r = self._v2("GET", "/pages", params={
            "spaceKey": space_key,
            "title": title,
            "limit": 1,
        })
        results = r.json().get("results", [])
        return results[0] if results else None

    def create_page(self, space_key: str, title: str, body_storage: str,
                    parent_id: str = "") -> dict:
        """
        Creates a new Confluence page with storage format body.
        Ref: POST /wiki/api/v2/pages
        """
        payload = {
            "spaceId": self._get_space_id(space_key),
            "status": "current",
            "title": title,
            "body": {
                "representation": "storage",
                "value": body_storage,
            },
        }
        if parent_id:
            payload["parentId"] = parent_id
        r = self._v2("POST", "/pages", json=payload)
        return r.json()

    def update_page(self, page_id: str, title: str, body_storage: str,
                    version: int) -> dict:
        """
        Updates an existing page body.
        Ref: PUT /wiki/api/v2/pages/{id}
        """
        payload = {
            "id": page_id,
            "status": "current",
            "title": title,
            "version": {"number": version + 1},
            "body": {
                "representation": "storage",
                "value": body_storage,
            },
        }
        r = self._v2("PUT", f"/pages/{page_id}", json=payload)
        return r.json()

    _space_id_cache: dict = {}

    def _get_space_id(self, space_key: str) -> str:
        """
        Resolves space key to space ID (required by v2 API).
        Ref: GET /wiki/api/v2/spaces
        """
        if space_key in self._space_id_cache:
            return self._space_id_cache[space_key]
        r = self._v2("GET", "/spaces", params={"keys": space_key, "limit": 1})
        results = r.json().get("results", [])
        if not results:
            raise ValueError(f"Space '{space_key}' not found in Confluence")
        sid = results[0]["id"]
        self._space_id_cache[space_key] = sid
        return sid

    def upload_attachment(self, page_id: str, filename: str,
                          data: bytes, mime_type: str) -> dict:
        """
        Uploads a file as an attachment to a Confluence page.
        Ref: POST /wiki/rest/api/content/{id}/child/attachment  (v1 — attachments
             are NOT in v2 API as of 2026; v2 returns 405 Method Not Allowed)
        Requires X-Atlassian-Token: no-check header to bypass XSRF protection.
        Uses multipart/form-data — Content-Type must NOT be set (requests sets it).
        """
        time.sleep(REQUEST_DELAY_S)
        url = f"{self.base}/wiki/rest/api/content/{page_id}/child/attachment"
        headers = {
            k: v for k, v in self.session.headers.items()
            if k.lower() not in ("content-type",)
        }
        headers["X-Atlassian-Token"] = "no-check"
        r = self.session.post(
            url,
            headers=headers,
            files={"file": (filename, data, mime_type)},
            data={"comment": "Migrated from Notion"},
        )
        if not r.ok:
            log.error(f"Attachment upload failed: {r.status_code} {r.text[:400]}")
        r.raise_for_status()
        return r.json()

    def apply_edit_restriction(self, page_id: str, admin_account_id: str) -> None:
        """
        Restricts 'update' operation on a page to a single admin user.
        This is the locking mechanism for notion_created_date.

        Ref: PUT /wiki/rest/api/content/{id}/restriction/byOperation/{operationKey}
        Correct payload per Confluence Cloud REST API docs (2024+):
          Array of user objects with type + accountId at the top level of results.
        Note: accountId must be the Atlassian Cloud accountId (the one from
          https://yoursite.atlassian.net/rest/api/3/myself — NOT a profile page ID).
          Run get_my_account_id() below to retrieve and verify your accountId.
        """
        if not admin_account_id:
            log.warning("  ⚠ CONFLUENCE_ADMIN_ACCOUNT_ID not set — skipping restriction")
            return

        # Correct endpoint: byOperation/{operationKey} with array body
        # Ref: https://developer.atlassian.com/cloud/confluence/rest/v1/api-group-content-restrictions/
        payload = [
            {
                "type": "known",
                "accountId": admin_account_id,
            }
        ]
        try:
            self._v1(
                "PUT",
                f"/content/{page_id}/restriction/byOperation/update/user",
                params={"accountId": admin_account_id},
                json=payload,
            )
            log.info(f"  ✓ Edit restriction applied to page {page_id}")
        except Exception as e:
            # Restriction failure is non-fatal — page content is already migrated
            log.warning(f"  ⚠ Edit restriction failed (non-fatal): {e}")
            log.warning(f"    Verify CONFLUENCE_ADMIN_ACCOUNT_ID via: "
                        f"curl -u email:token {self.base}/wiki/rest/api/user/current")

    def get_my_account_id(self) -> str:
        """
        Helper to retrieve your Atlassian accountId — use this to verify
        CONFLUENCE_ADMIN_ACCOUNT_ID is correct before running restrictions.
        Prints the correct value to use in your .env file.
        Ref: GET /wiki/rest/api/user/current
        """
        r = self._v1("GET", "/user/current")
        account_id = r.json().get("accountId", "")
        log.info(f"Your Confluence accountId: {account_id}")
        log.info(f"Set CONFLUENCE_ADMIN_ACCOUNT_ID={account_id} in your .env file")
        return account_id

    def get_page_version(self, page_id: str) -> int:
        """Returns current version number of a page."""
        r = self._v2("GET", f"/pages/{page_id}")
        return r.json().get("version", {}).get("number", 1)


class JiraClient:
    """
    Wraps Jira Cloud REST API v3.
    Ref: https://developer.atlassian.com/cloud/jira/platform/rest/v3/
    Used only to write the Confluence backlink into customfield_10085.
    """
    def __init__(self, base_url: str, email: str, token: str):
        self.base = base_url.rstrip("/")
        creds = base64.b64encode(f"{email}:{token}".encode()).decode()
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization": f"Basic {creds}",
            "Content-Type": "application/json",
            "Accept": "application/json",
        })

    def update_issue_confluence_link(self, issue_key: str,
                                     confluence_url: str) -> None:
        """
        Writes Confluence page URL to customfield_10085 on a Jira issue.
        Ref: PUT /rest/api/3/issue/{issueIdOrKey}
        """
        time.sleep(REQUEST_DELAY_S)
        url = f"{self.base}/rest/api/3/issue/{issue_key}"
        payload = {"fields": {JIRA_CONFLUENCE_FIELD: confluence_url}}
        r = self.session.put(url, json=payload)
        if not r.ok:
            log.warning(f"Jira update failed for {issue_key}: {r.status_code} {r.text[:200]}")
        else:
            log.info(f"  ✓ Jira {issue_key} → Confluence URL written")


# ─────────────────────────────────────────────
# NOTION BLOCK → CONFLUENCE STORAGE FORMAT
# ─────────────────────────────────────────────

def _escape_xml(text: str) -> str:
    """Escape special XML characters for Confluence storage format."""
    return (text
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;"))


def _rich_text_to_storage(rich_texts: list) -> str:
    """
    Converts Notion rich_text array to Confluence storage format HTML.
    Handles: bold, italic, strikethrough, underline, code, color, links.
    Ref: https://developer.atlassian.com/cloud/confluence/confluence-storage-format/
    """
    out = ""
    for rt in rich_texts:
        text = _escape_xml(rt.get("plain_text", ""))
        ann = rt.get("annotations", {})
        href = rt.get("href")

        if ann.get("code"):
            text = f"<code>{text}</code>"
        if ann.get("bold"):
            text = f"<strong>{text}</strong>"
        if ann.get("italic"):
            text = f"<em>{text}</em>"
        if ann.get("strikethrough"):
            text = f"<s>{text}</s>"
        if ann.get("underline"):
            text = f"<u>{text}</u>"
        color = ann.get("color", "default")
        if color and color != "default" and not color.endswith("_background"):
            # Confluence storage supports ac:parameter color via structured macros;
            # inline color applied via span for plain rich text
            text = f'<span style="color:{color};">{text}</span>'
        if href:
            text = f'<a href="{_escape_xml(href)}">{text}</a>'
        out += text
    return out


def blocks_to_storage(blocks: list, confluence_client=None,
                      page_id: str = "", attachment_map: dict = None) -> str:
    """
    Recursively converts Notion block list to Confluence storage format XML.
    Handles:
      paragraph, heading_1/2/3, bulleted_list_item, numbered_list_item,
      to_do, toggle, quote, code, divider, callout, table, table_row,
      image, file, pdf, video, embed, bookmark, child_page, synced_block,
      column_list, column.
    Attachments (image/file/pdf): downloaded from Notion and uploaded to
    Confluence if confluence_client and page_id are provided.
    attachment_map: dict to accumulate {notion_url: confluence_attachment_url}
    """
    if attachment_map is None:
        attachment_map = {}

    html = ""
    # Track list state for wrapping <ul>/<ol>
    prev_list_type = None
    list_buffer = ""

    def flush_list():
        nonlocal prev_list_type, list_buffer, html
        if list_buffer:
            tag = "ul" if prev_list_type == "ul" else "ol"
            html += f"<{tag}>{list_buffer}</{tag}>"
            list_buffer = ""
            prev_list_type = None

    for block in blocks:
        btype = block.get("type", "")
        # Guard: Notion occasionally returns null block content
        # e.g. {"type":"paragraph","paragraph":null}
        data = block.get(btype) or {}
        children_html = ""
        if block.get("_children"):
            try:
                children_html = blocks_to_storage(
                    block["_children"], confluence_client, page_id, attachment_map
                )
            except Exception as _ce:
                log.warning(f"  ⚠ Child blocks failed [{btype}]: {_ce} — skipping children")
                children_html = ""

        # ── List items ──────────────────────────────────────────────────────
        if btype == "bulleted_list_item":
            if prev_list_type != "ul":
                flush_list()
                prev_list_type = "ul"
            inner = _rich_text_to_storage(data.get("rich_text", []))
            if children_html:
                inner += f"<ul>{children_html}</ul>"
            list_buffer += f"<li>{inner}</li>"
            continue

        if btype == "numbered_list_item":
            if prev_list_type != "ol":
                flush_list()
                prev_list_type = "ol"
            inner = _rich_text_to_storage(data.get("rich_text", []))
            if children_html:
                inner += f"<ol>{children_html}</ol>"
            list_buffer += f"<li>{inner}</li>"
            continue

        # Any non-list block flushes the pending list
        flush_list()

        # One bad block must never crash the whole page conversion
        try:
            # ── Headings ─────────────────────────────────────────────────────────
            if btype == "heading_1":
                html += f"<h1>{_rich_text_to_storage(data.get('rich_text', []))}</h1>"

            elif btype == "heading_2":
                html += f"<h2>{_rich_text_to_storage(data.get('rich_text', []))}</h2>"

            elif btype == "heading_3":
                html += f"<h3>{_rich_text_to_storage(data.get('rich_text', []))}</h3>"

            # ── Paragraph ────────────────────────────────────────────────────────
            elif btype == "paragraph":
                inner = _rich_text_to_storage(data.get("rich_text", []))
                if not inner and not children_html:
                    html += "<p>&nbsp;</p>"
                else:
                    html += f"<p>{inner}{children_html}</p>"

            # ── To-do ────────────────────────────────────────────────────────────
            elif btype == "to_do":
                checked = data.get("checked", False)
                check = "✅ " if checked else "☐ "
                inner = _rich_text_to_storage(data.get("rich_text", []))
                html += f"<p>{check}{inner}</p>"

            # ── Quote ────────────────────────────────────────────────────────────
            elif btype == "quote":
                inner = _rich_text_to_storage(data.get("rich_text", []))
                html += (
                    f'<blockquote><p>{inner}</p>'
                    f'{children_html}</blockquote>'
                )

            # ── Code ─────────────────────────────────────────────────────────────
            elif btype == "code":
                lang = data.get("language", "plain text")
                code_text = _escape_xml(
                    "".join(rt.get("plain_text", "")
                            for rt in data.get("rich_text", []))
                )
                # Confluence Code Block macro
                html += (
                    f'<ac:structured-macro ac:name="code" ac:schema-version="1">'
                    f'<ac:parameter ac:name="language">{_escape_xml(lang)}</ac:parameter>'
                    f'<ac:plain-text-body><![CDATA[{code_text}]]></ac:plain-text-body>'
                    f'</ac:structured-macro>'
                )

            # ── Divider ──────────────────────────────────────────────────────────
            elif btype == "divider":
                html += "<hr/>"

            # ── Callout ──────────────────────────────────────────────────────────
            elif btype == "callout":
                inner = _rich_text_to_storage(data.get("rich_text", []))
                emoji = ""
                icon = data.get("icon", {})
                if icon.get("type") == "emoji":
                    emoji = icon.get("emoji", "")
                html += (
                    f'<ac:structured-macro ac:name="info" ac:schema-version="1">'
                    f'<ac:rich-text-body>'
                    f'<p><strong>{emoji}</strong> {inner}</p>'
                    f'{children_html}'
                    f'</ac:rich-text-body>'
                    f'</ac:structured-macro>'
                )

            # ── Toggle ───────────────────────────────────────────────────────────
            elif btype == "toggle":
                inner = _rich_text_to_storage(data.get("rich_text", []))
                html += (
                    f'<ac:structured-macro ac:name="expand" ac:schema-version="1">'
                    f'<ac:parameter ac:name="title">{inner}</ac:parameter>'
                    f'<ac:rich-text-body>{children_html}</ac:rich-text-body>'
                    f'</ac:structured-macro>'
                )

            # ── Table ────────────────────────────────────────────────────────────
            elif btype == "table":
                has_header = data.get("has_column_header", False)
                table_html = "<table><tbody>"
                for i, row_block in enumerate(block.get("_children", [])):
                    cells = row_block.get("table_row", {}).get("cells", [])
                    table_html += "<tr>"
                    for cell in cells:
                        cell_content = _rich_text_to_storage(cell)
                        tag = "th" if (has_header and i == 0) else "td"
                        table_html += f"<{tag}>{cell_content}</{tag}>"
                    table_html += "</tr>"
                table_html += "</tbody></table>"
                html += table_html
                continue  # children already processed above

            elif btype == "table_row":
                continue  # handled inside table block

            # ── Column list / Column ─────────────────────────────────────────────
            elif btype == "column_list":
                cols = block.get("_children", [])
                col_count = max(len(cols), 1)
                pct = 100 // col_count
                layout_html = (
                    f'<ac:structured-macro ac:name="column-layout" ac:schema-version="1">'
                    f'<ac:rich-text-body>'
                )
                for col in cols:
                    col_inner = blocks_to_storage(
                        col.get("_children", []), confluence_client,
                        page_id, attachment_map
                    )
                    layout_html += (
                        f'<ac:structured-macro ac:name="column" ac:schema-version="1">'
                        f'<ac:parameter ac:name="width">{pct}%</ac:parameter>'
                        f'<ac:rich-text-body>{col_inner}</ac:rich-text-body>'
                        f'</ac:structured-macro>'
                    )
                layout_html += f'</ac:rich-text-body></ac:structured-macro>'
                html += layout_html
                continue

            elif btype == "column":
                continue  # handled inside column_list

            # ── Image ────────────────────────────────────────────────────────────
            elif btype == "image":
                img_url, filename = _extract_file_url(data, "image")
                if img_url:
                    att_html = _handle_file_attachment(
                        img_url, filename, confluence_client, page_id,
                        attachment_map, is_image=True
                    )
                    caption = _rich_text_to_storage(data.get("caption", []))
                    html += att_html
                    if caption:
                        html += f"<p><em>{caption}</em></p>"

            # ── File (generic, includes embedded docs) ───────────────────────────
            elif btype == "file":
                file_url, filename = _extract_file_url(data, "file")
                if file_url:
                    att_html = _handle_file_attachment(
                        file_url, filename, confluence_client, page_id,
                        attachment_map, is_image=False
                    )
                    caption = _rich_text_to_storage(data.get("caption", []))
                    html += att_html
                    if caption:
                        html += f"<p><em>{caption}</em></p>"

            # ── PDF ──────────────────────────────────────────────────────────────
            elif btype == "pdf":
                pdf_url, filename = _extract_file_url(data, "pdf")
                if pdf_url:
                    # Upload as attachment, then embed via PDF macro
                    att_html = _handle_file_attachment(
                        pdf_url, filename, confluence_client, page_id,
                        attachment_map, is_image=False, is_pdf=True
                    )
                    html += att_html

            # ── Bookmark / Link Preview ──────────────────────────────────────────
            elif btype in ("bookmark", "link_preview"):
                url = data.get("url", "")
                caption = _rich_text_to_storage(data.get("caption", []))
                display = caption if caption else _escape_xml(url)
                if url:
                    html += f'<p><a href="{_escape_xml(url)}">{display}</a></p>'

            # ── Embed (YouTube, PPTX preview, Figma, etc.) ───────────────────────
            elif btype == "embed":
                url = data.get("url", "")
                if url:
                    # Use Confluence Widget Connector macro for embeddable URLs
                    html += (
                        f'<ac:structured-macro ac:name="widget" ac:schema-version="1">'
                        f'<ac:parameter ac:name="url">{_escape_xml(url)}</ac:parameter>'
                        f'</ac:structured-macro>'
                    )

            # ── Video ────────────────────────────────────────────────────────────
            elif btype == "video":
                vid_url, _ = _extract_file_url(data, "video")
                if vid_url:
                    html += (
                        f'<ac:structured-macro ac:name="widget" ac:schema-version="1">'
                        f'<ac:parameter ac:name="url">{_escape_xml(vid_url)}</ac:parameter>'
                        f'</ac:structured-macro>'
                    )

            # ── Child page reference ──────────────────────────────────────────────
            elif btype == "child_page":
                title = data.get("title", "")
                child_id = block.get("id", "")
                html += (
                    f'<p>📄 <em>Child page: {_escape_xml(title)}'
                    f' (Notion ID: {child_id})</em></p>'
                )

            # ── Synced block ──────────────────────────────────────────────────────
            elif btype == "synced_block":
                if children_html:
                    html += children_html
                else:
                    # Synced from another block — children are the content
                    html += f"<p><em>[Synced block — content above]</em></p>"

            # ── Equation ─────────────────────────────────────────────────────────
            elif btype == "equation":
                expr = _escape_xml(data.get("expression", ""))
                html += f"<p><code>{expr}</code></p>"

            # ── Unknown block type — preserve as note ────────────────────────────
            else:
                if btype not in ("unsupported",):
                    log.debug(f"Unhandled block type: {btype}")
                    html += f"<p><em>[Block type '{btype}' not rendered]</em></p>"

        except Exception as _block_err:
            block_id = block.get("id", "unknown")
            log.warning(f"  ⚠ Block render error [type={btype!r} id={block_id!r}]: {_block_err}")
            html += (
                "<ac:structured-macro ac:name=\"warning\" ac:schema-version=\"1\">"
                "<ac:rich-text-body>"
                f"<p>&#9888; Block type <code>{btype}</code> "
                f"(Notion id: {block_id}) could not be rendered. "
                f"Error: <code>{_escape_xml(str(_block_err))}</code></p>"
                "</ac:rich-text-body>"
                "</ac:structured-macro>"
            )
    flush_list()
    return html


def _extract_file_url(data: dict, block_type: str) -> tuple[str, str]:
    """
    Extracts URL and filename from a Notion file/image/pdf block.
    Returns (url, filename).
    Handles both 'external' (public URL) and 'file' (signed S3 URL) types.
    """
    file_obj = data.get("external") or data.get("file")
    if not file_obj:
        return "", f"{block_type}_file"
    url = file_obj.get("url", "")
    # Extract filename from URL path (before query string)
    path_part = url.split("?")[0]
    filename = path_part.split("/")[-1] or f"{block_type}_file"
    # Ensure extension
    if "." not in filename:
        ext_map = {"image": ".png", "pdf": ".pdf", "file": ".bin", "video": ".mp4"}
        filename += ext_map.get(block_type, ".bin")
    return url, filename


def _handle_file_attachment(url: str, filename: str, confluence_client,
                             page_id: str, attachment_map: dict,
                             is_image: bool = False,
                             is_pdf: bool = False) -> str:
    """
    Downloads a file from Notion, uploads to Confluence as an attachment,
    and returns Confluence storage XML to inline or link it.

    For images: uses ac:image macro.
    For PDFs: uses viewfile macro (renders inline PDF viewer).
    For other files: uses ac:link to attachment.
    """
    if url in attachment_map:
        return attachment_map[url]

    if not confluence_client or not page_id:
        # Dry run or no client — emit placeholder link
        result = f'<p><a href="{_escape_xml(url)}">{_escape_xml(filename)}</a></p>'
        attachment_map[url] = result
        return result

    try:
        notion_client_temp = NotionClient.__new__(NotionClient)
        notion_client_temp.session = requests.Session()
        file_bytes = requests.get(url, timeout=60).content
        mime, _ = mimetypes.guess_type(filename)
        mime = mime or "application/octet-stream"

        att_result = confluence_client.upload_attachment(
            page_id, filename, file_bytes, mime
        )
        log.info(f"  ✓ Attachment uploaded: {filename}")

        safe_name = _escape_xml(filename)
        if is_image:
            result = (
                f'<ac:image ac:thumbnail="false">'
                f'<ri:attachment ri:filename="{safe_name}"/>'
                f'</ac:image>'
            )
        elif is_pdf:
            result = (
                f'<ac:structured-macro ac:name="viewfile" ac:schema-version="1">'
                f'<ac:parameter ac:name="name">{safe_name}</ac:parameter>'
                f'</ac:structured-macro>'
            )
        else:
            result = (
                f'<ac:link>'
                f'<ri:attachment ri:filename="{safe_name}"/>'
                f'<ac:plain-text-link-body><![CDATA[{filename}]]></ac:plain-text-link-body>'
                f'</ac:link>'
            )
        attachment_map[url] = result
        return result

    except Exception as e:
        log.warning(f"  ⚠ Attachment failed [{filename}]: {e}")
        result = f'<p><a href="{_escape_xml(url)}">{_escape_xml(filename)}</a></p>'
        attachment_map[url] = result
        return result


# ─────────────────────────────────────────────
# PAGE PROPERTIES MACRO BUILDER
# ─────────────────────────────────────────────

def build_page_properties_macro(props: dict, page_id_notion: str) -> str:
    """
    Builds a Confluence Page Properties macro (storage format XML) containing
    all migrated properties as a two-column table.

    The macro enables:
      - Page Properties Report across the space
      - Structured metadata visible on every page

    Ref: https://confluence.atlassian.com/doc/page-properties-macro-184550024.html
    Storage format: ac:structured-macro name="details"
    """
    rows = ""

    def add_row(key: str, value: str) -> str:
        safe_key = _escape_xml(key)
        safe_val = _escape_xml(str(value)) if value else ""
        return (
            f"<tr>"
            f"<th><p>{safe_key}</p></th>"
            f"<td><p>{safe_val}</p></td>"
            f"</tr>"
        )

    # notion_created_date — always first, labeled as write-once
    created = props.get("Created", "")
    if created:
        # Normalize to date-only (drop time component)
        created_display = created[:10] if len(created) >= 10 else created
    else:
        created_display = ""
    rows += add_row("notion_created_date", created_display)
    rows += add_row("notion_page_id", props.get("_notion_page_id", ""))

    # Ordered remaining properties
    ordered_keys = [
        ("status",        "Status"),
        ("priority",      "Priority"),
        ("category",      "Category"),
        ("artifact_type", "Artifact Type"),
        ("tags",          "Tags"),
        ("company",       "Company"),
        ("project",       "Project"),
        ("entry_date",    "Date"),
        ("author",        "Author"),
        ("source_url",    "URL"),
        ("ai_summary",    "AI Summary"),
        ("jira_issue_link", "Jira Issue"),
    ]
    for macro_key, notion_key in ordered_keys:
        val = props.get(notion_key, "")
        if val:
            rows += add_row(macro_key, val)

    macro = (
        f'<ac:structured-macro ac:name="details" ac:schema-version="1">'
        f'<ac:rich-text-body>'
        f'<table><tbody>{rows}</tbody></table>'
        f'</ac:rich-text-body>'
        f'</ac:structured-macro>'
        f'<hr/>'
    )
    return macro


# ─────────────────────────────────────────────
# BUILD SHEET (XLSX manifest)
# ─────────────────────────────────────────────

SHEET_COLS = [
    "Notion Page ID",
    "Notion Title",
    "Notion URL",
    "Notion Created",
    "Status",
    "Priority",
    "Category",
    "Artifact Type",
    "Tags",
    "Company",
    "Project",
    "Entry Date",
    "Author",
    "Source URL",
    "AI Summary (truncated)",
    "Has Attachments",
    "Confluence Page ID",
    "Confluence URL",
    "Jira Issue Key",
    "Migration Status",
    "Migrated At",
    "Error",
]

HDR_FILL  = PatternFill("solid", start_color="1F4E79")
HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
DONE_FILL = PatternFill("solid", start_color="E2EFDA")
ERR_FILL  = PatternFill("solid", start_color="FCE4D6")
SKIP_FILL = PatternFill("solid", start_color="EDEDED")
WRAP      = Alignment(wrap_text=True, vertical="top")
THIN      = Side(border_style="thin", color="D9D9D9")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_WIDTHS = {
    "A": 38, "B": 45, "C": 55, "D": 20, "E": 14, "F": 12,
    "G": 20, "H": 22, "I": 35, "J": 14, "K": 22, "L": 14,
    "M": 20, "N": 45, "O": 30, "P": 14, "Q": 18, "R": 55,
    "S": 16, "T": 16, "U": 20, "V": 40,
}


def create_build_sheet(pages_props: list) -> None:
    """
    Creates the dry-run build sheet XLSX with one row per Notion page.
    If the file already exists, loads it and only adds NEW rows
    (idempotent — never duplicates already-present page IDs).
    """
    existing_ids = set()
    if Path(BUILD_SHEET_PATH).exists():
        wb = load_workbook(BUILD_SHEET_PATH)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                existing_ids.add(str(row[0]))
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Migration Manifest"

        # Header row
        for col, name in enumerate(SHEET_COLS, 1):
            c = ws.cell(row=1, column=col, value=name)
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = WRAP
            c.border = BORDER
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28

        for col_letter, width in COL_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width

    added = 0
    for props in pages_props:
        pid = props.get("_notion_page_id", "")
        if pid in existing_ids:
            continue

        # Detect attachments
        has_att = bool(
            props.get("File Attachments", "").strip() or
            props.get("Media Attachments", "").strip()
        )

        row_data = [
            pid,
            props.get("Name", ""),
            props.get("_notion_url", ""),
            (props.get("Created", "") or "")[:10],
            props.get("Status", ""),
            props.get("Priority", ""),
            props.get("Category", ""),
            props.get("Artifact Type", ""),
            props.get("Tags", ""),
            props.get("Company", ""),
            props.get("Project", ""),
            (props.get("Date", "") or "")[:10],
            props.get("Author", ""),
            props.get("URL", ""),
            (props.get("AI summary", "") or "")[:120],
            "YES" if has_att else "",
            "",   # Confluence Page ID — filled after migration
            "",   # Confluence URL
            "",   # Jira Issue Key
            "PENDING",
            "",   # Migrated At
            "",   # Error
        ]

        next_row = ws.max_row + 1
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=next_row, column=col, value=val)
            c.font = BODY_FONT
            c.alignment = WRAP
            c.border = BORDER
        added += 1

    wb.save(BUILD_SHEET_PATH)
    log.info(f"Build sheet saved: {BUILD_SHEET_PATH} ({added} new rows added)")


def update_build_sheet_row(notion_page_id: str, confluence_page_id: str,
                            confluence_url: str, jira_key: str,
                            status: str, error: str = "") -> None:
    """
    Updates a single row in the build sheet after a page is migrated.
    Finds the row by Notion Page ID in column A.
    """
    if not Path(BUILD_SHEET_PATH).exists():
        log.warning("Build sheet not found — cannot update row")
        return

    wb = load_workbook(BUILD_SHEET_PATH)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == notion_page_id:
            row[16].value = confluence_page_id   # Q: Confluence Page ID
            row[17].value = confluence_url        # R: Confluence URL
            row[18].value = jira_key              # S: Jira Issue Key
            row[19].value = status                # T: Migration Status
            row[20].value = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
            row[21].value = error                 # V: Error

            fill = DONE_FILL if status == "DONE" else (
                ERR_FILL if status == "ERROR" else SKIP_FILL
            )
            for cell in row:
                cell.fill = fill
            break

    wb.save(BUILD_SHEET_PATH)


# ─────────────────────────────────────────────
# MIGRATION ORCHESTRATOR
# ─────────────────────────────────────────────

def run_phase_0_build_sheet(notion: NotionClient) -> list:
    """Phase 0: Query Notion, extract all properties, write build sheet."""
    log.info("=" * 60)
    log.info("PHASE 0 — Building dry-run manifest")
    log.info("=" * 60)
    pages = notion.query_database(NOTION_DB_ID)
    all_props = [notion.extract_properties(p) for p in pages]
    create_build_sheet(all_props)
    log.info(f"Phase 0 complete — {len(all_props)} pages in manifest")
    return all_props


def run_migration(notion: NotionClient, confluence: ConfluenceClient,
                  jira: JiraClient, admin_account_id: str,
                  all_props: list = None) -> None:
    """
    Phases 1–4: For each page in the build sheet with status PENDING:
      1. Retrieve blocks from Notion
      2. Convert to Confluence storage format
      3. Prepend Page Properties macro
      4. Create Confluence page
      5. Upload attachments
      6. Apply edit restriction on notion_created_date
      7. Write Confluence URL back to Jira
      8. Mark row as DONE in build sheet
    """
    if DRY_RUN:
        log.info("DRY_RUN=True — skipping live writes. Set DRY_RUN=False to migrate.")
        return

    if all_props is None:
        # Reload from build sheet
        wb = load_workbook(BUILD_SHEET_PATH)
        ws = wb.active
        all_props = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[19] == "PENDING":
                all_props.append({
                    "_notion_page_id": row[0],
                    "Name": row[1],
                    "_notion_url": row[2],
                    "Created": row[3],
                    "Status": row[4],
                    "Priority": row[5],
                    "Category": row[6],
                    "Artifact Type": row[7],
                    "Tags": row[8],
                    "Company": row[9],
                    "Project": row[10],
                    "Date": row[11],
                    "Author": row[12],
                    "URL": row[13],
                    "AI summary": row[14],
                    "_jira_issue_link": "",
                })

    # ── Apply batch slicing ──────────────────────────────────────────────────
    if BATCH_SIZE > 0:
        start = BATCH_NUMBER * BATCH_SIZE
        end   = start + BATCH_SIZE
        batch = all_props[start:end]
        log.info("=" * 60)
        log.info(f"PHASE 1–4 — Batch {BATCH_NUMBER}: pages {start+1}–{min(end, len(all_props))} "
                 f"of {len(all_props)} PENDING")
        log.info(f"  Next run: set BATCH_NUMBER = {BATCH_NUMBER + 1}")
        log.info("=" * 60)
    else:
        batch = all_props
        log.info("=" * 60)
        log.info(f"PHASE 1–4 — Migrating ALL {len(all_props)} PENDING pages (BATCH_SIZE=0)")
        log.info("=" * 60)

    if not batch:
        log.info("No pages in this batch range — all pages may already be migrated.")
        log.info(f"Total PENDING pages: {len(all_props)}. "
                 f"Try lowering BATCH_NUMBER or check build sheet.")
        return

    parent_id = CONFLUENCE_PARENT_ID or ""

    for i, props in enumerate(batch, 1):
        pid = props.get("_notion_page_id", "")
        title = props.get("Name", f"Untitled_{pid[:8]}")
        log.info(f"[{i}/{len(batch)}] Processing: {title}")

        try:
            # ── Phase 1: Check for existing page ────────────────────────────
            existing = confluence.find_page_by_title(CONFLUENCE_SPACE_KEY, title)
            if existing:
                log.info(f"  → Already exists in Confluence, skipping create")
                cf_page_id = existing["id"]
                cf_url = f"{confluence.base}/wiki/spaces/{CONFLUENCE_SPACE_KEY}/pages/{cf_page_id}"
                update_build_sheet_row(pid, cf_page_id, cf_url, "", "SKIPPED (exists)")
                continue

            # ── Phase 2: Fetch Notion blocks ─────────────────────────────────
            log.info(f"  → Fetching blocks from Notion...")
            blocks = notion.get_page_blocks(pid)

            # ── Phase 3: Convert blocks → storage format ──────────────────────
            # First pass: create the page with properties macro only
            # (need page ID to upload attachments before second pass)
            props_macro = build_page_properties_macro(props, pid)
            placeholder_body = props_macro + "<p><em>Content loading...</em></p>"

            # ── Phase 4: Create Confluence page ──────────────────────────────
            log.info(f"  → Creating Confluence page...")
            created_page = confluence.create_page(
                CONFLUENCE_SPACE_KEY, title, placeholder_body, parent_id
            )
            cf_page_id = created_page["id"]
            cf_version = created_page.get("version", {}).get("number", 1)
            cf_url = f"{confluence.base}/wiki/spaces/{CONFLUENCE_SPACE_KEY}/pages/{cf_page_id}"
            log.info(f"  ✓ Page created: {cf_url}")

            # ── Phase 5: Convert blocks with real attachment uploads ───────────
            log.info(f"  → Converting {len(blocks)} blocks...")
            attachment_map = {}
            # blocks_to_storage is now fault-tolerant — bad blocks render as
            # warning macros instead of raising, so this should always return HTML
            content_html = blocks_to_storage(
                blocks, confluence, cf_page_id, attachment_map
            )

            # Inject Confluence URL into properties for backlink display
            props["_jira_issue_link"] = props.get("_jira_issue_link", "")
            props_macro = build_page_properties_macro(props, pid)
            final_body = props_macro + content_html

            # ── Phase 5b: Update page with real content ───────────────────────
            # Always runs — even if some blocks had warnings, the rest of the
            # content is there. Page will never be left showing "Content loading..."
            log.info(f"  → Updating page with full content ({len(content_html)} chars)...")
            # Re-fetch version number in case Confluence incremented it
            current_version = confluence.get_page_version(cf_page_id)
            confluence.update_page(cf_page_id, title, final_body, current_version)
            log.info(f"  ✓ Content written to Confluence page")

            # ── Phase 6: Lock notion_created_date via edit restriction ─────────
            # Non-fatal: restriction failure does not roll back the migrated page
            if admin_account_id:
                log.info(f"  → Applying edit restriction...")
                confluence.apply_edit_restriction(cf_page_id, admin_account_id)

            # ── Phase 7: Write Confluence URL back to Jira ────────────────────
            jira_key = props.get("_jira_issue_link", "")
            if jira_key:
                log.info(f"  → Updating Jira issue {jira_key}...")
                try:
                    jira.update_issue_confluence_link(jira_key, cf_url)
                except Exception as je:
                    log.warning(f"  ⚠ Jira update non-fatal: {je}")

            # ── Phase 8: Mark as DONE ─────────────────────────────────────────
            # Page content is successfully migrated at this point regardless of
            # whether restriction or Jira backlink succeeded
            update_build_sheet_row(pid, cf_page_id, cf_url, jira_key, "DONE")
            log.info(f"  ✓ DONE: {title}")

        except Exception as e:
            err_msg = f"{type(e).__name__}: {str(e)[:200]}"
            log.error(f"  ✗ FAILED [{title}]: {err_msg}")
            log.debug(traceback.format_exc())
            # If page was created before the error, record its ID so it isn't
            # re-created on retry (idempotent: find_page_by_title will skip it)
            partial_id = locals().get("cf_page_id", "")
            partial_url = locals().get("cf_url", "")
            update_build_sheet_row(pid, partial_id, partial_url, "", "ERROR", err_msg)


# ─────────────────────────────────────────────
# CREDENTIAL VALIDATION
# ─────────────────────────────────────────────

def validate_credentials() -> dict:
    """
    Validates all required environment variables are set.
    Does NOT make API calls — purely checks env presence.
    Returns dict of {name: value} for all creds.
    """
    required = {
        "NOTION_TOKEN"         : "Notion integration secret",
        "CONFLUENCE_BASE_URL"  : "e.g. https://yoursite.atlassian.net",
        "CONFLUENCE_EMAIL"     : "Atlassian account email",
        "CONFLUENCE_API_TOKEN" : "Atlassian API token",
        "JIRA_BASE_URL"        : "e.g. https://yoursite.atlassian.net",
        "JIRA_EMAIL"           : "Jira account email",
        "JIRA_API_TOKEN"       : "Jira API token",
        "CONFLUENCE_SPACE_KEY" : "Target Confluence space key",
        "CONFLUENCE_ADMIN_ACCOUNT_ID": "Atlassian accountId for restriction lock",
    }
    missing = []
    creds = {}
    for key, desc in required.items():
        val = os.getenv(key, "")
        if not val:
            missing.append(f"  {key}  ({desc})")
        creds[key] = val

    if missing:
        log.error("Missing required environment variables:")
        for m in missing:
            log.error(m)
        if not DRY_RUN:
            sys.exit(1)
        else:
            log.warning("DRY_RUN=True — continuing without live credentials")
    return creds


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────

def main():
    log.info("=" * 60)
    log.info("Notion → Confluence Migration")
    log.info(f"Mode         : {'DRY RUN — no writes' if DRY_RUN else 'LIVE'}")
    log.info(f"Source DB    : 06. Work ({NOTION_DB_ID})")
    log.info(f"Target Space : {CONFLUENCE_SPACE_KEY}")
    if not DRY_RUN:
        batch_desc = f"ALL pages" if BATCH_SIZE == 0 else f"Batch {BATCH_NUMBER} ({BATCH_SIZE} pages)"
        log.info(f"Batch        : {batch_desc}")
    log.info("=" * 60)

    creds = validate_credentials()

    notion = NotionClient(creds.get("NOTION_TOKEN", "placeholder"))

    if not DRY_RUN:
        confluence = ConfluenceClient(
            creds["CONFLUENCE_BASE_URL"],
            creds["CONFLUENCE_EMAIL"],
            creds["CONFLUENCE_API_TOKEN"],
        )
        jira = JiraClient(
            creds["JIRA_BASE_URL"],
            creds["JIRA_EMAIL"],
            creds["JIRA_API_TOKEN"],
        )
        admin_id = creds.get("CONFLUENCE_ADMIN_ACCOUNT_ID", "")
    else:
        confluence = None
        jira = None
        admin_id = ""

    # Verify accountId on first live run (helps catch wrong CONFLUENCE_ADMIN_ACCOUNT_ID)
    if not DRY_RUN and confluence:
        log.info("Verifying Confluence account ID...")
        verified_id = confluence.get_my_account_id()
        if verified_id and verified_id != creds.get("CONFLUENCE_ADMIN_ACCOUNT_ID", ""):
            log.warning("⚠ CONFLUENCE_ADMIN_ACCOUNT_ID in .env does not match your actual")
            log.warning(f"  accountId. Update .env: CONFLUENCE_ADMIN_ACCOUNT_ID={verified_id}")
            admin_id = verified_id   # auto-correct for this run

    # Phase 0 — always runs (build sheet is safe even in dry run)
    all_props = run_phase_0_build_sheet(notion)

    # Phases 1–4 — only runs when DRY_RUN=False
    if not DRY_RUN:
        run_migration(notion, confluence, jira, admin_id, all_props)

    log.info("=" * 60)
    log.info(f"Run complete. Build sheet: {BUILD_SHEET_PATH}")
    if DRY_RUN:
        log.info("Next steps:")
        log.info("  1. Open notion_confluence_build_sheet.xlsx and review all rows")
        log.info("  2. Set DRY_RUN = False")
        log.info("  3. Set BATCH_SIZE = 5 and BATCH_NUMBER = 0")
        log.info("  4. Run again — 5 pages will migrate to Confluence")
        log.info("  5. Check those 5 pages in Confluence, then increment BATCH_NUMBER")
    else:
        if BATCH_SIZE > 0:
            log.info(f"Batch {BATCH_NUMBER} done. To continue: set BATCH_NUMBER = {BATCH_NUMBER + 1} and re-run.")
        else:
            log.info("All pages processed. Check build sheet for any ERROR rows.")
    log.info("=" * 60)


if __name__ == "__main__":
    main()