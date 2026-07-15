import requests
from requests.auth import HTTPBasicAuth
from openpyxl import Workbook
from openpyxl.styles import Font

# ==================== CONFIGURATION ====================
DOMAIN = "https://YOUR-DOMAIN.atlassian.net"
EMAIL = "your-email@example.com"
API_TOKEN = "your-api-token-here"

# Set up Auth and Headers
auth = HTTPBasicAuth(EMAIL, API_TOKEN)
headers = {"Accept": "application/json"}

def fetch_data(endpoint):
    """Helper function to make the API GET request."""
    url = f"{DOMAIN}{endpoint}"
    print(f"Fetching: {endpoint}...")
    response = requests.get(url, headers=headers, auth=auth)
    
    if response.status_code == 200:
        return response.json()
    else:
        print(f"❌ Failed to fetch {endpoint} - Status: {response.status_code}")
        return None

def write_sheet(ws, title, headers, data_rows):
    """Helper function to write data and format headers in Excel."""
    ws.title = title
    ws.append(headers)
    # Bold the header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    for row in data_rows:
        ws.append(row)

# ==================== MAIN SCRIPT ====================
def main():
    print("🚀 Starting Jira Configuration Export...")
    wb = Workbook()
    
    # 1. FETCH FIELDS (Custom & System)
    fields_data = fetch_data("/rest/api/3/field")
    ws_fields = wb.active # Use default first sheet
    if fields_data:
        parsed_fields = [[f.get("id"), f.get("name"), f.get("custom", False)] for f in fields_data]
        write_sheet(ws_fields, "Fields", ["Field ID", "Field Name", "Is Custom Field?"], parsed_fields)

    # 2. FETCH PROJECTS
    projects_data = fetch_data("/rest/api/3/project/search?expand=description,lead")
    if projects_data and "values" in projects_data:
        ws_proj = wb.create_sheet()
        parsed_projects = [[
            p.get("key"), 
            p.get("name"), 
            p.get("projectTypeKey"), 
            p.get("lead", {}).get("displayName", "N/A")
        ] for p in projects_data["values"]]
        write_sheet(ws_proj, "Projects", ["Project Key", "Name", "Type", "Project Lead"], parsed_projects)

    # 3. FETCH WORKFLOWS
    workflows_data = fetch_data("/rest/api/3/workflow/search")
    if workflows_data and "values" in workflows_data:
        ws_wf = wb.create_sheet()
        parsed_workflows = [[
            w.get("id", {}).get("name"), 
            w.get("description", ""), 
            len(w.get("statuses", []))
        ] for w in workflows_data["values"]]
        write_sheet(ws_wf, "Workflows", ["Workflow Name", "Description", "Total Statuses"], parsed_workflows)

    # 4. FETCH ISSUE TYPES
    issue_types_data = fetch_data("/rest/api/3/issuetype")
    if issue_types_data:
        ws_it = wb.create_sheet()
        parsed_types = [[
            it.get("id"), 
            it.get("name"), 
            it.get("description", ""), 
            it.get("subtask", False)
        ] for it in issue_types_data]
        write_sheet(ws_it, "Issue Types", ["ID", "Name", "Description", "Is Subtask?"], parsed_types)

    # 5. FETCH SYSTEM SETTINGS (Application Properties)
    app_props_data = fetch_data("/rest/api/3/application-properties")
    if app_props_data:
        ws_sys = wb.create_sheet()
        parsed_props = [[
            p.get("id"), 
            p.get("name"), 
            p.get("value")
        ] for p in app_props_data]
        write_sheet(ws_sys, "System Settings", ["Property ID", "Name", "Value"], parsed_props)

    # ==================== SAVE FILE ====================
    filename = "Jira_Configuration_Export.xlsx"
    wb.save(filename)
    print(f"\n✅ Export complete! Saved to {filename} in your current directory.")

if __name__ == "__main__":
    main()