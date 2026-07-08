#!/usr/bin/env python3
"""
Notion Block Type Scanner
Version: 1.0

Scans all 8 Mission Control Notion databases, walks every page's block tree,
counts every block type used, and outputs:
  1. XLSX mapping table — block type, count, pages found in, recommended
     Confluence equivalent, implementation status, notes
  2. Terminal summary — sorted by frequency

Usage:
  python3 notion_block_scanner.py

Requires .env with NOTION_TOKEN set.
"""

import os
import sys
import time
import logging
from collections import defaultdict
from dotenv import load_dotenv
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

NOTION_TOKEN   = os.getenv("NOTION_TOKEN", "")
REQUEST_DELAY  = 0.2   # seconds between API calls
OUTPUT_PATH    = "notion_block_mapping.xlsx"
MAX_PAGES      = 0     # 0 = all pages; set to e.g. 20 to sample

# All 8 Mission Control databases
# Name → Notion database ID
DATABASES = {
    "01. Career Development" : "37761870-08a2-811f-a3d7-000ca47d0ec7",
    "02. Learning"           : "37761870-08a2-819a-a709-000b1ee1c40f",
    "03. Family & Life"      : "37761870-08a2-8116-a725-000b136c7722",
    "04. Tech & Tinkering"   : "37761870-08a2-81bb-b279-000be6527130",
    "05. News & Articles"    : "37761870-08a2-810f-b932-000bcdae4f04",
    "06. Work"               : "37761870-08a2-8199-941a-c6e481ccf03e",
    "07. Divorce"            : "37761870-08a2-8133-a656-000b8f1fbe4d",
    "08. Faith & Spiritual"  : "37761870-08a2-81ad-b250-000bbea73438",
}
# ─────────────────────────────────────────────
# CONFLUENCE MAPPING TABLE
# Notion block type → recommended Confluence storage format element
# Status: supported / partial / unsupported
# ─────────────────────────────────────────────
CONFLUENCE_MAPPING = {
    "paragraph": {
        "confluence": "<p>...</p>",
        "macro": None,
        "status": "✅ Supported",
        "notes": "Direct HTML paragraph. Works in Fabric.",
    },
    "heading_1": {
        "confluence": "<h1>...</h1>",
        "macro": None,
        "status": "✅ Supported",
        "notes": "Direct HTML heading. Works in Fabric.",
    },
    "heading_2": {
        "confluence": "<h2>...</h2>",
        "macro": None,
        "status": "✅ Supported",
        "notes": "Direct HTML heading. Works in Fabric.",
    },
    "heading_3": {
        "confluence": "<h3>...</h3>",
        "macro": None,
        "status": "✅ Supported",
        "notes": "Direct HTML heading. Works in Fabric.",
    },
    "bulleted_list_item": {
        "confluence": "<ul><li>...</li></ul>",
        "macro": None,
        "status": "✅ Supported",
        "notes": "Standard HTML list. Works in Fabric.",
    },
    "numbered_list_item": {
        "confluence": "<ol><li>...</li></ol>",
        "macro": None,
        "status": "✅ Supported",
        "notes": "Standard HTML list. Works in Fabric.",
    },
    "to_do": {
        "confluence": "<p>☐ / ✅ text</p>",
        "macro": None,
        "status": "⚠️ Partial",
        "notes": "Rendered as emoji + text paragraph. No native checkbox in storage format.",
    },
    "toggle": {
        "confluence": "ac:name=\"expand\"",
        "macro": "expand",
        "status": "✅ Supported",
        "notes": "Expand macro. Works in Fabric — verified safe.",
    },
    "quote": {
        "confluence": "<blockquote><p>...</p></blockquote>",
        "macro": None,
        "status": "✅ Supported",
        "notes": "Standard HTML blockquote. Works in Fabric.",
    },
    "divider": {
        "confluence": "<hr/>",
        "macro": None,
        "status": "✅ Supported",
        "notes": "Standard HTML horizontal rule. Works in Fabric.",
    },
    "code": {
        "confluence": "ac:name=\"code\"",
        "macro": "code",
        "status": "✅ Supported",
        "notes": "Confluence Code Block macro. Works in Fabric.",
    },
    "callout": {
        "confluence": "ac:name=\"info\"",
        "macro": "info",
        "status": "⚠️ Partial",
        "notes": "Info macro loses emoji/color variants. KNOWN ISSUE: causes legacy content warning in Fabric when wrapped in expand macro. Use info macro ONLY — do not nest in expand.",
    },
    "table": {
        "confluence": "<table><tbody>...</tbody></table>",
        "macro": None,
        "status": "✅ Supported",
        "notes": "Standard HTML table with th/td. Works in Fabric.",
    },
    "table_row": {
        "confluence": "<tr><td>...</td></tr>",
        "macro": None,
        "status": "✅ Supported",
        "notes": "Handled inside table block.",
    },
    "column_list": {
        "confluence": "ac:name=\"column-layout\"",
        "macro": "column-layout",
        "status": "⚠️ Partial",
        "notes": "KNOWN ISSUE: column-layout macro causes legacy content warning in Fabric. Flatten columns to sequential paragraphs instead.",
    },
    "column": {
        "confluence": "ac:name=\"column\"",
        "macro": "column",
        "status": "⚠️ Partial",
        "notes": "Child of column_list. Same legacy warning issue. Flatten content.",
    },
    "image": {
        "confluence": "<ac:image><ri:attachment/></ac:image>",
        "macro": None,
        "status": "✅ Supported",
        "notes": "Upload as attachment then reference with ac:image tag.",
    },
    "file": {
        "confluence": "<ac:link><ri:attachment/></ac:link>",
        "macro": None,
        "status": "✅ Supported",
        "notes": "Upload as attachment then link with ac:link tag.",
    },
    "pdf": {
        "confluence": "ac:name=\"viewfile\"",
        "macro": "viewfile",
        "status": "✅ Supported",
        "notes": "Upload as attachment then embed with viewfile macro.",
    },
    "video": {
        "confluence": "ac:name=\"widget\"",
        "macro": "widget",
        "status": "⚠️ Partial",
        "notes": "Widget Connector macro for YouTube/Vimeo URLs. Local video files not supported.",
    },
    "embed": {
        "confluence": "ac:name=\"widget\"",
        "macro": "widget",
        "status": "⚠️ Partial",
        "notes": "Widget Connector macro. Works for YouTube/Vimeo/public URLs only.",
    },
    "bookmark": {
        "confluence": "<p><a href=\"...\">...</a></p>",
        "macro": None,
        "status": "✅ Supported",
        "notes": "Rendered as hyperlink paragraph. No rich preview.",
    },
    "link_preview": {
        "confluence": "<p><a href=\"...\">...</a></p>",
        "macro": None,
        "status": "✅ Supported",
        "notes": "Same as bookmark — hyperlink only, no embed preview.",
    },
    "child_page": {
        "confluence": "<p>📄 reference text</p>",
        "macro": None,
        "status": "⚠️ Partial",
        "notes": "Rendered as a text reference. Child page hierarchy not migrated.",
    },
    "synced_block": {
        "confluence": "Inline content",
        "macro": None,
        "status": "⚠️ Partial",
        "notes": "Children rendered inline. Sync relationship not preserved.",
    },
    "equation": {
        "confluence": "<p><code>expression</code></p>",
        "macro": None,
        "status": "⚠️ Partial",
        "notes": "Rendered as code text. No LaTeX rendering in Confluence storage format.",
    },
    "unsupported": {
        "confluence": "ac:name=\"warning\" (placeholder)",
        "macro": "warning",
        "status": "❌ Not supported",
        "notes": "Notion returns 'unsupported' for Tabs blocks and other non-API block types. Content inaccessible via API. Warning macro with link back to Notion page.",
    },
    "table_of_contents": {
        "confluence": "ac:name=\"toc\"",
        "macro": "toc",
        "status": "✅ Supported",
        "notes": "Confluence Table of Contents macro.",
    },
    "breadcrumb": {
        "confluence": "ac:name=\"breadcrumbs\"",
        "macro": "breadcrumbs",
        "status": "✅ Supported",
        "notes": "Confluence Breadcrumbs macro.",
    },
    "child_database": {
        "confluence": "<p>reference text</p>",
        "macro": None,
        "status": "❌ Not supported",
        "notes": "Inline databases are Notion-specific. Rendered as text reference only.",
    },
    "template": {
        "confluence": None,
        "macro": None,
        "status": "❌ Not supported",
        "notes": "Notion template buttons have no Confluence equivalent.",
    },
    "link_to_page": {
        "confluence": "<p><a href=\"...\">...</a></p>",
        "macro": None,
        "status": "⚠️ Partial",
        "notes": "Rendered as hyperlink. Target page may not exist in Confluence.",
    },
}

# ─────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# NOTION API
# ─────────────────────────────────────────────

BASE = "https://api.notion.com/v1"
HEADERS = {
    "Authorization": f"Bearer {NOTION_TOKEN}",
    "Notion-Version": "2022-06-28",
    "Content-Type": "application/json",
}

def notion_get(path: str, params: dict = None) -> dict:
    time.sleep(REQUEST_DELAY)
    r = requests.get(f"{BASE}{path}", headers=HEADERS, params=params)
    r.raise_for_status()
    return r.json()

def notion_post(path: str, body: dict) -> dict:
    time.sleep(REQUEST_DELAY)
    r = requests.post(f"{BASE}{path}", headers=HEADERS, json=body)
    r.raise_for_status()
    return r.json()

def query_database(db_id: str) -> list:
    results, cursor = [], None
    while True:
        body = {"page_size": 100}
        if cursor:
            body["start_cursor"] = cursor
        data = notion_post(f"/databases/{db_id}/query", body)
        results.extend(data.get("results", []))
        if not data.get("has_more"):
            break
        cursor = data.get("next_cursor")
    return results

def get_blocks_recursive(block_id: str, depth: int = 0) -> list:
    """Recursively fetch all blocks. Returns flat list of (block_type, depth, block_id)."""
    found = []
    cursor = None
    while True:
        params = {"page_size": 100}
        if cursor:
            params["start_cursor"] = cursor
        try:
            data = notion_get(f"/blocks/{block_id}/children", params=params)
        except Exception as e:
            log.warning(f"  Block fetch error for {block_id}: {e}")
            break
        blocks = data.get("results", [])
        for block in blocks:
            btype = block.get("type", "unknown")
            bid   = block.get("id", "")
            found.append((btype, depth, bid))
            if block.get("has_children") and btype not in ("unsupported",):
                found.extend(get_blocks_recursive(bid, depth + 1))
        if not data.get("has_more"):
            break
        cursor = data.get("next_cursor")
    return found

# ─────────────────────────────────────────────
# SCANNER
# ─────────────────────────────────────────────

def scan_all_databases() -> tuple[dict, dict, dict]:
    """
    Returns:
      block_counts  : {block_type: total_count}
      block_pages   : {block_type: set of page titles}
      block_dbs     : {block_type: set of db names}
    """
    block_counts = defaultdict(int)
    block_pages  = defaultdict(set)
    block_dbs    = defaultdict(set)

    total_pages_scanned = 0

    for db_name, db_id in DATABASES.items():
        log.info(f"\n{'='*55}")
        log.info(f"Scanning: {db_name}")
        log.info(f"{'='*55}")

        try:
            pages = query_database(db_id)
        except Exception as e:
            log.error(f"  Failed to query {db_name}: {e}")
            continue

        if MAX_PAGES > 0:
            pages = pages[:MAX_PAGES]

        log.info(f"  {len(pages)} pages found")

        for idx, page in enumerate(pages, 1):
            page_id    = page.get("id", "")
            props      = page.get("properties", {})
            # Get title
            title = ""
            for _, prop in props.items():
                if prop.get("type") == "title":
                    title = "".join(
                        rt.get("plain_text", "")
                        for rt in prop.get("title", [])
                    )
                    break
            if not title:
                title = f"Untitled ({page_id[:8]})"

            log.info(f"  [{idx}/{len(pages)}] {title[:60]}")

            try:
                blocks = get_blocks_recursive(page_id)
            except Exception as e:
                log.warning(f"    Block scan error: {e}")
                continue

            for btype, depth, _ in blocks:
                block_counts[btype] += 1
                block_pages[btype].add(title[:50])
                block_dbs[btype].add(db_name)

            total_pages_scanned += 1

    log.info(f"\n{'='*55}")
    log.info(f"Scan complete — {total_pages_scanned} pages scanned")
    log.info(f"{'='*55}\n")
    return block_counts, block_pages, block_dbs

# ─────────────────────────────────────────────
# TERMINAL SUMMARY
# ─────────────────────────────────────────────

def print_terminal_summary(block_counts: dict, block_pages: dict, block_dbs: dict):
    sorted_blocks = sorted(block_counts.items(), key=lambda x: -x[1])
    total_blocks  = sum(block_counts.values())

    print(f"\n{'='*70}")
    print(f"  NOTION BLOCK TYPE INVENTORY — {len(sorted_blocks)} unique types, {total_blocks:,} total blocks")
    print(f"{'='*70}")
    print(f"  {'BLOCK TYPE':<28} {'COUNT':>7}  {'%':>5}  {'STATUS':<20}  DBs")
    print(f"  {'-'*28} {'-'*7}  {'-'*5}  {'-'*20}  {'-'*20}")

    for btype, count in sorted_blocks:
        mapping = CONFLUENCE_MAPPING.get(btype, {})
        status  = mapping.get("status", "❓ Unknown")
        pct     = count / total_blocks * 100
        dbs     = ", ".join(sorted(block_dbs[btype]))[:30]
        print(f"  {btype:<28} {count:>7,}  {pct:>4.1f}%  {status:<20}  {dbs}")

    print(f"\n  Total blocks scanned: {total_blocks:,}")
    print(f"  Unique block types:   {len(sorted_blocks)}")
    known_unsupported = [b for b in block_counts if
                         CONFLUENCE_MAPPING.get(b, {}).get("status", "").startswith("❌")]
    known_partial     = [b for b in block_counts if
                         CONFLUENCE_MAPPING.get(b, {}).get("status", "").startswith("⚠️")]
    print(f"  ❌ Not supported:     {len(known_unsupported)} types — {', '.join(known_unsupported)}")
    print(f"  ⚠️  Partial support:   {len(known_partial)} types — {', '.join(known_partial)}")
    print(f"{'='*70}\n")

# ─────────────────────────────────────────────
# XLSX OUTPUT
# ─────────────────────────────────────────────

HDR_BLUE   = PatternFill("solid", start_color="1F4E79")
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT  = Font(name="Arial", size=10)
GREEN_FILL = PatternFill("solid", start_color="E2EFDA")
AMBER_FILL = PatternFill("solid", start_color="FFF2CC")
RED_FILL   = PatternFill("solid", start_color="FCE4D6")
GREY_FILL  = PatternFill("solid", start_color="F2F2F2")
WRAP       = Alignment(wrap_text=True, vertical="top")
THIN       = Side(border_style="thin", color="D9D9D9")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def build_xlsx(block_counts: dict, block_pages: dict, block_dbs: dict):
    wb = Workbook()

    # ── Sheet 1: Block Mapping ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Block Mapping"

    headers = [
        "Notion Block Type",
        "Count",
        "% of All Blocks",
        "Confluence Element / Macro",
        "Macro Name",
        "Status",
        "Implementation Notes",
        "Databases Found In",
        "Sample Pages (first 3)",
    ]
    col_widths = [28, 10, 12, 30, 18, 18, 55, 40, 60]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HDR_FONT
        c.fill = HDR_BLUE
        c.alignment = WRAP
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    total_blocks = sum(block_counts.values())
    sorted_blocks = sorted(block_counts.items(), key=lambda x: -x[1])

    for row_idx, (btype, count) in enumerate(sorted_blocks, 2):
        mapping = CONFLUENCE_MAPPING.get(btype, {})
        status  = mapping.get("status", "❓ Unknown — add mapping")
        pct     = count / total_blocks * 100
        dbs     = "\n".join(sorted(block_dbs[btype]))
        pages   = "\n".join(list(block_pages[btype])[:3])

        row_data = [
            btype,
            count,
            f"{pct:.1f}%",
            mapping.get("confluence", "— NOT MAPPED —"),
            mapping.get("macro", ""),
            status,
            mapping.get("notes", "No mapping defined — needs implementation"),
            dbs,
            pages,
        ]

        if status.startswith("✅"):
            fill = GREEN_FILL
        elif status.startswith("⚠️"):
            fill = AMBER_FILL
        elif status.startswith("❌"):
            fill = RED_FILL
        else:
            fill = GREY_FILL

        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.font = BODY_FONT
            c.alignment = WRAP
            c.border = BORDER
            if col in (1, 6):
                c.fill = fill

        ws.row_dimensions[row_idx].height = max(
            30, 15 * (1 + dbs.count("\n"))
        )

    # ── Sheet 2: Action Items ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Action Items")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 70

    action_headers = ["Notion Block Type", "Status", "Required Action"]
    for col, h in enumerate(action_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = HDR_FONT
        c.fill = HDR_BLUE
        c.alignment = WRAP
        c.border = BORDER

    action_row = 2
    actions = {
        "callout": (
            "⚠️ Partial",
            "REMOVE from expand macro wrapper — use info macro directly at top level. "
            "The expand+info nesting causes 'This is legacy content' warning in Fabric. "
            "Render as: ac:name='info' at block level, NOT nested inside ac:name='expand'."
        ),
        "column_list": (
            "⚠️ Partial",
            "REPLACE column-layout macro with sequential paragraphs. "
            "column-layout macro causes 'This is legacy content' warning in Fabric. "
            "Flatten each column's content in order: left column first, then right column."
        ),
        "column": (
            "⚠️ Partial",
            "Child of column_list — same fix. Render children inline, not in column macro."
        ),
        "unsupported": (
            "❌ Not supported",
            "Notion Tabs block — content inaccessible via API. "
            "Render as warning macro with link back to original Notion page."
        ),
        "to_do": (
            "⚠️ Partial",
            "No native checkbox in Confluence storage format. "
            "Current implementation (emoji + text) is acceptable. No action needed."
        ),
    }

    for btype, (status, action) in actions.items():
        if btype in block_counts:
            ws2.cell(row=action_row, column=1, value=btype).font = BODY_FONT
            ws2.cell(row=action_row, column=2, value=status).font = BODY_FONT
            c = ws2.cell(row=action_row, column=3, value=action)
            c.font = BODY_FONT
            c.alignment = WRAP
            ws2.row_dimensions[action_row].height = 60
            action_row += 1

    # ── Sheet 3: Summary Stats ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20

    stats = [
        ("Total blocks scanned", sum(block_counts.values())),
        ("Unique block types found", len(block_counts)),
        ("Databases scanned", len(DATABASES)),
        ("", ""),
        ("✅ Fully supported", len([b for b in block_counts if
            CONFLUENCE_MAPPING.get(b, {}).get("status", "").startswith("✅")])),
        ("⚠️  Partial support", len([b for b in block_counts if
            CONFLUENCE_MAPPING.get(b, {}).get("status", "").startswith("⚠️")])),
        ("❌ Not supported", len([b for b in block_counts if
            CONFLUENCE_MAPPING.get(b, {}).get("status", "").startswith("❌")])),
        ("❓ Unknown / unmapped", len([b for b in block_counts if
            b not in CONFLUENCE_MAPPING])),
        ("", ""),
        ("KNOWN FABRIC ISSUES", ""),
        ("callout macro", "Causes legacy warning when nested in expand — use flat info macro"),
        ("column-layout macro", "Causes legacy warning — flatten to sequential paragraphs"),
        ("expand macro (toggle)", "Safe to use at top level — verified working"),
        ("details macro (Page Properties)", "Causes legacy warning — remove from output"),
        ("expand wrapping details", "Causes legacy warning — this was the root cause of the Fabric editor issue"),
    ]

    for row_idx, (label, value) in enumerate(stats, 1):
        c1 = ws3.cell(row=row_idx, column=1, value=label)
        c2 = ws3.cell(row=row_idx, column=2, value=value)
        c1.font = Font(name="Arial", size=10, bold=(label.startswith(("✅","⚠️","❌","❓","KNOWN","Total","Unique","Databases"))))
        c2.font = Font(name="Arial", size=10)
        c1.alignment = WRAP
        c2.alignment = WRAP

    wb.save(OUTPUT_PATH)
    log.info(f"XLSX saved: {OUTPUT_PATH}")

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    if not NOTION_TOKEN:
        log.error("NOTION_TOKEN not set in .env — cannot connect to Notion API")
        sys.exit(1)

    log.info("Notion Block Type Scanner v1.0")
    log.info(f"Scanning {len(DATABASES)} databases")
    if MAX_PAGES > 0:
        log.info(f"Sample mode: first {MAX_PAGES} pages per database")
    log.info("")

    block_counts, block_pages, block_dbs = scan_all_databases()

    if not block_counts:
        log.error("No blocks found — check database IDs and NOTION_TOKEN")
        sys.exit(1)

    print_terminal_summary(block_counts, block_pages, block_dbs)
    build_xlsx(block_counts, block_pages, block_dbs)

    print(f"\nOutput: {OUTPUT_PATH}")
    print("Next step: review Block Mapping and Action Items sheets,")
    print("then update notion_to_confluence.py using the mapping table.")

if __name__ == "__main__":
    main()